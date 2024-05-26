import openpyxl

class PlanilhaVendas:
    def __init__(self, vendedores):
        self.vendedores = vendedores
        self.planilha = openpyxl.load_workbook(self.vendedores)
        self.pagina_vendedor = self.planilha['Plan1']
        
    def dados(self):
        for linha in self.pagina_vendedor.iter_rows(min_row=2):
            print(f'Data venda: {linha[0].value}, \nNome vendedor: {linha[1].value}, \nValor Venda: {linha[2].value}, \nTipo Cliente: {linha[3].value},\nCanal de venda: {linha[4].value}\n')

    def total_vendas(self):
        total = 0
        for linha in self.pagina_vendedor.iter_rows(min_row=2):
            valor_venda = linha[2].value
            if valor_venda is not None:
                total += valor_venda
        return "{:,.2f}".format(total).replace(",", "X").replace(".", ",").replace("X", ".")
       
        
  
planilha_vendedores = PlanilhaVendas('Planilha_vendas.xlsx')
planilha_vendedores.dados()
print(f'Total de vendas: R${planilha_vendedores.total_vendas()}')



