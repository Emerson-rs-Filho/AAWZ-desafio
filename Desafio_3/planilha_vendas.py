import openpyxl

class PlanilhaVendas:
    def __init__(self, vendedores):
        self.vendedores = vendedores
        self.planilha = openpyxl.load_workbook(self.vendedores)
        self.pagina_vendedor = self.planilha['Plan1']
        
    def dados(self):
        for linha in self.pagina_vendedor.iter_rows(min_row=2):
            print(f'Data venda: {linha[0].value}, \nNome vendedor: {linha[1].value}, \nValor Venda: {linha[2].value}, \nTipo Cliente: {linha[3].value},\n Canal de venda: {linha[4].value}\n')

    def calc_comicoes(self):
        comissoes = {}

        for linha in self.pagina_vendedor.iter_rows(min_row=2):
            vendedor = linha[1].value
            valor_venda = linha[2].value
            canal_venda = linha[4].value

            if vendedor not in comissoes:
                comissoes[vendedor] = {'comissao': 0, 'marketing': 0, 'gerente': 0}
            
            comissao_vendedor = valor_venda * 0.10
            comissoes[vendedor]['comissao'] += comissao_vendedor
            
            if canal_venda.lower() == 'online':
                comissao_marketing = valor_venda * 0.20
                comissoes[vendedor]['marketing'] += comissao_marketing

            if valor_venda >= 1000:
                comissao_gerente = valor_venda * 0.10
                comissoes[vendedor]['gerente'] += comissao_gerente

        return comissoes

planilha_vendedores = PlanilhaVendas('Planilha_vendas.xlsx')
planilha_vendedores.dados()

comissoes = planilha_vendedores.calc_comicoes()
for vendedor, valores in comissoes.items():
    print(f'Vendedor: {vendedor}, \nComissão Final: {valores["comissao"]}, \nComissão para Marketing: {valores["marketing"]}, \nComissão para Gerente: {valores["gerente"]}\n')




