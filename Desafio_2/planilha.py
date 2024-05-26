import openpyxl
########################################################################################
#Cpf existente na planilha para teste de atualização de dados
# - 567.890.123-54
# - 456.789.012-43
# - 345.678.901-32
########################################################################################

class PlanilhaVendedores:
    def __init__(self, vendedores):
        self.vendedores = vendedores
        self.planilha = openpyxl.load_workbook(self.vendedores)
        self.pagina_vendedor = self.planilha['Vendedores']
        
    def nome_vendedores(self):
        for linha in self.pagina_vendedor.iter_rows(min_row=2):
            print(f'Nome: {linha[0].value}, \nCPF: {linha[1].value}, \nData nascimento: {linha[2].value}, \nEstado: {linha[3].value}\n')

    def alterar_dados(self):
        
        cpf = input("Digite o CPF do vendedor que deseja alterar: ")
        
        cpf_existe = False
        for linha in self.pagina_vendedor.iter_rows(min_row=2):
            if linha[1].value == cpf:
                cpf_existe = True
                break

        if cpf_existe:    
            novo_nome = input("Digite o novo nome: ")
            nova_data_nascimento = input("Digite a nova data de nascimento: ")
            novo_estado = input("Digite o novo estado: ")

            for linha in self.pagina_vendedor.iter_rows(min_row=2):
                if linha[1].value == cpf:
                    linha[0].value = novo_nome
                    linha[2].value = nova_data_nascimento
                    linha[3].value = novo_estado
                    break

            self.planilha.save(self.vendedores)
            print("Dados atualizados com sucesso.")
        else:
            print("CPF não encontrado na planilha.")
            return

planilha_vendedores = PlanilhaVendedores('AAWZ_Vendedores.xlsx')
planilha_vendedores.alterar_dados()
planilha_vendedores.nome_vendedores()